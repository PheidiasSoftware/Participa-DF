from datetime import datetime
from pathlib import Path
import argparse

from joblib import dump
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.load_data import load_data
from src.preprocess import apply_weak_labels
from src.model import train_and_evaluate

# Define caminhos padrao para dados
DEFAULT_DATA = "data/dataset_treinamento_acesso_informacao.xlsx"
FALLBACK_DATA = "data/AMOSTRA_e-SIC.xlsx"
OUTPUT_CSV = "outputs/rotulos_fracos.csv"
OUTPUT_XLSX = "outputs/rotulos_fracos.xlsx"
METRICS_PATH = "reports/metrics.txt"


def save_xlsx_with_widths(df, path):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(path)

def main():
    # CLI para escolher planilha de entrada
    parser = argparse.ArgumentParser(
        description=(
            "Treina e avalia o modelo de classificacao para dados pessoais. "
            "Suporta datasets com rotulos ou rotulagem fraca."
        )
    )
    parser.add_argument(
        "--data",
        default=DEFAULT_DATA if Path(DEFAULT_DATA).exists() else FALLBACK_DATA,
        help=(
            "Caminho do arquivo XLSX. "
            "Pode conter colunas ID/Texto Mascarado ou id/text/label."
        ),
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=0.65,
        help=(
            "Limiar minimo para manter non_public sem evidencias de PII. "
            "Valor e salvo em models/threshold.txt."
        ),
    )
    parser.add_argument(
        "--test-rules",
        action="store_true",
        help="Executa validacao das regras em data/casos_dificeis.csv antes do treino.",
    )
    args = parser.parse_args()

    if args.test_rules:
        try:
            from avancado import test_rules
            test_rules.main()
        except Exception as exc:
            print(f"Erro ao executar test_rules: {exc}")

    try:
        # Carrega dados em formato padrao (id, text, label opcional)
        df = load_data(args.data)
    except ValueError as exc:
        print(f"Erro ao carregar dados: {exc}")
        return
    if "label" in df.columns:
        # Normaliza rótulos supervisionados
        df["label"] = df["label"].astype(str).str.strip().str.lower()
    else:
        # Aplica rotulagem fraca quando o dataset nao tem label
        df = apply_weak_labels(df)
    if "weight" not in df.columns:
        # Peso padrao para rotulos supervisionados
        df["weight"] = 1.0
    # Salva dataset rotulado para reproducao
    Path(OUTPUT_CSV).parent.mkdir(exist_ok=True)
    df.to_csv(OUTPUT_CSV, index=False, sep=";", encoding="utf-8-sig")
    save_xlsx_with_widths(df, OUTPUT_XLSX)
    report_text, model, vectorizer = train_and_evaluate(df)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Salva relatorio de metricas com data/hora
    Path(METRICS_PATH).parent.mkdir(exist_ok=True)
    with open(METRICS_PATH, "w", encoding="utf-8") as f:
        f.write(f"Data/Hora: {timestamp}\n")
        f.write(report_text)

    model_dir = Path("models")
    # Garante pasta de modelos
    model_dir.mkdir(exist_ok=True)
    # Persiste vetor e modelo
    dump(vectorizer, model_dir / "vectorizer.joblib")
    dump(model, model_dir / "model.joblib")

    # Salva limiar usado para referencia na inferencia
    with open(model_dir / "threshold.txt", "w", encoding="utf-8") as f:
        f.write(str(args.threshold))

if __name__ == "__main__":
    main()
