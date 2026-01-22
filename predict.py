import argparse
from pathlib import Path

import pandas as pd
from joblib import load
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.rules import has_personal_data


def load_input(path: Path) -> pd.DataFrame:
    # Le XLSX ou CSV e normaliza colunas
    if path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)

    # Renomeia para o padrao interno
    if "Texto Mascarado" in df.columns:
        df = df.rename(columns={"Texto Mascarado": "text", "ID": "id"})
    elif "text" in df.columns:
        if "id" not in df.columns:
            # Gera ids sequenciais se nao existirem
            df["id"] = range(1, len(df) + 1)
    else:
        raise ValueError("Entrada deve conter coluna 'Texto Mascarado' ou 'text'.")

    # Mantem apenas as colunas usadas na predicao
    df = df[["id", "text"]]
    df["text"] = df["text"].astype(str).str.strip()
    return df


def save_xlsx_with_widths(df, path: Path):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(path)


def main():
    # CLI para predicao por texto unico ou por arquivo
    parser = argparse.ArgumentParser(
        description="Classifica pedidos usando o modelo salvo em models/."
    )
    parser.add_argument(
        "--text",
        help="Texto unico para classificar (alternativa ao --input).",
    )
    parser.add_argument(
        "--input",
        help="Arquivo .xlsx/.csv com coluna Texto Mascarado ou text.",
    )
    parser.add_argument(
        "--output",
        default="outputs/predictions.csv",
        help="CSV de saida com labels e scores.",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=None,
        help=(
            "Limiar minimo para manter non_public sem evidencias de PII. "
            "Se nao informado, usa models/threshold.txt."
        ),
    )
    args = parser.parse_args()

    # Limiar minimo para manter non_public sem evidencias de PII
    if args.threshold is not None:
        threshold_non_public = args.threshold
    else:
        try:
            with open("models/threshold.txt", "r", encoding="utf-8") as f:
                threshold_non_public = float(f.read().strip())
        except Exception:
            threshold_non_public = 0.65

    # Carrega modelo e vetor salvos
    model = load("models/model.joblib")
    vectorizer = load("models/vectorizer.joblib")

    if args.text:
        # Predicao para texto unico
        X_vec = vectorizer.transform([args.text])
        pred = model.predict(X_vec)[0]
        rule_hit = has_personal_data(args.text)
        score = None
        if hasattr(model, "predict_proba"):
            score = model.predict_proba(X_vec)[0].max()
        # Regra tem prioridade para evitar falsos negativos
        if rule_hit:
            final_label = "non_public"
        else:
            final_label = pred
            # Evita falsos positivos quando nao ha evidencias explicitas
            if pred == "non_public" and score is not None and score < threshold_non_public:
                final_label = "public"

        if score is not None:
            print(f"label_modelo={pred} label_final={final_label} score={score:.4f} regra_pii={rule_hit}")
        else:
            print(f"label_modelo={pred} label_final={final_label} regra_pii={rule_hit}")
        return

    if not args.input:
        raise SystemExit("Use --text ou --input.")

    # Predicao em lote por arquivo
    df = load_input(Path(args.input))
    X_vec = vectorizer.transform(df["text"])
    df["label_modelo"] = model.predict(X_vec)
    df["regra_pii"] = df["text"].apply(has_personal_data)
    if hasattr(model, "predict_proba"):
        df["score"] = model.predict_proba(X_vec).max(axis=1)
        df["label"] = df["label_modelo"]
        # Regra tem prioridade para evitar falsos negativos
        df.loc[df["regra_pii"] == True, "label"] = "non_public"
        mask = (
            (df["label_modelo"] == "non_public")
            & (~df["regra_pii"])
            & (df["score"] < threshold_non_public)
        )
        df.loc[mask, "label"] = "public"
    else:
        df["label"] = df["label_modelo"]
        df.loc[df["regra_pii"] == True, "label"] = "non_public"

    # Salva resultados
    out_path = Path(args.output)
    out_path.parent.mkdir(exist_ok=True)
    df.to_csv(out_path, index=False, sep=";", encoding="utf-8-sig")

    xlsx_path = out_path.with_suffix(".xlsx")
    save_xlsx_with_widths(df, xlsx_path)

    print(f"Arquivo gerado: {out_path}")
    print(f"Arquivo XLSX gerado: {xlsx_path}")


if __name__ == "__main__":
    main()
